from openpyxl import load_workbook

excelfile = load_workbook(filename='Sales-2024-07-27.xlsx')

sheet_list = excelfile.sheetnames

total = 0

for st in sheet_list:
    # st = sheet name : 'somchai'
    sheet = excelfile[st]
    print(sheet['A2'].value)
    total = total + sheet['A2'].value

print(total)