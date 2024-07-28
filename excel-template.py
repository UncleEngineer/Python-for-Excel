# excel-template.py

from openpyxl import load_workbook

data_excel = load_workbook(filename='data-invoice.xlsx')
sheet_data = data_excel.active # data_excel['ใบกำกับภาษี']

# template_excel = load_workbook(filename='template-iv.xlsx')
# sheet_template = template_excel.active # template_excel[0]

# print(sheet_data['E2'].value)

# number = sheet_data['E2'].value
# # แก้ไข cell ใน template
# sheet_template['S14'] = number
# template_excel.save('result-excel.xlsx')

count = len(sheet_data['A']) # มีจำนวนกี่แถว
print(count)

rows = [] # สร้างลิสต์เพื่อเก็บข้อมูลแถว นำ row ทั้งมาใส่ในนี้

for i in range(2,count+1):
    datalist = [] # แถว 1 แถว
    for d in sheet_data[i]:
        datalist.append(d.value) # d.value คือ column
        #print(datalist)
    rows.append(datalist)
        #print('-----')

print(len(rows))

# loop เพื่อสร้างกลุ่ม IV ว่ามีเลข IV กี่ตัว
rows_dict = {}
for r in rows:
    if r[4] not in rows_dict:
        # r[4] = IV xxxxxxxx เลข transaction
        rows_dict[r[4]] = [] # สร้าง key เป็นรายการไว้

# print(rows_dict)

# loop เพื่อแยก IV แต่ละ transaction
for r in rows:
    rows_dict[r[4]].append(r)

###################นำ row dict ไปหยอดใน template####################
######RECORD 1#######
count_name = 1 # นับจำนวนไฟล์

for rw in rows_dict.values():
    # rw = [['บริษัท A',....],['บริษัท A',....],['บริษัท A',....]]
    template_excel = load_workbook(filename='template-iv.xlsx')
    sheet_template = template_excel.active # template_excel[0]
    row_product = 25
    # state เอาไว้เช็ค
    check_company = False
    check_address1 = False
    check_address2 = False
    check_taxid = False
    check_number = False
    check_date = False # ****เพิ่มใหม่****
    for i,r in enumerate(rw,start=1):
        data =  r # rows_dict['IV-6307021'][0]
        company = data[0]
        address1 = data[1]
        address2 = data[2]
        taxid = data[3]
        number = data[4]
        date = data[9] # ****เพิ่มใหม่****
        if check_company == False:
            sheet_template['C13'] = company
            check_company = True
        if check_address1 == False:
            sheet_template['C15'] = address1
            check_address1 = True
        if check_address2 == False:
            sheet_template['C17'] = address2
            check_address2 = True
        if check_taxid == False:
            sheet_template['D20'] = taxid
            check_taxid = True
        if check_number == False:
            sheet_template['S14'] = number
            check_number = True
        if check_date == False:
            sheet_template['S16'] = date # ****เพิ่มใหม่****

        # ข้อมูลสินค้า 1 บรรทัด
        product_no = data[5]
        product_item = data[6]
        product_quantity = data[7]
        product_price = data[8]

        # 3,5,11,17
        sheet_template.cell(row=row_product,column=2).value = i
        sheet_template.cell(row=row_product,column=3).value = product_no
        sheet_template.cell(row=row_product,column=5).value = product_item
        sheet_template.cell(row=row_product,column=11).value = product_quantity
        sheet_template.cell(row=row_product,column=14).value = 'PCS.'
        sheet_template.cell(row=row_product,column=17).value = product_price
        row_product = row_product + 1
    
    template_excel.save('result-data-template-{}.xlsx'.format(number))
    count_name += 1 # count_name = count_name + 1

